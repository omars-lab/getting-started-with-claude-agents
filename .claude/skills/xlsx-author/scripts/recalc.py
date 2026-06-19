#!/usr/bin/env python3
"""
Recalculate an .xlsx workbook using the pure-Python `formulas` library and
report any formula errors as JSON.

Usage:
    .venv/bin/python3 .claude/skills/xlsx-author/scripts/recalc.py path/to/workbook.xlsx

Why this exists: openpyxl writes formula *strings* but does not evaluate them.
A workbook can save successfully and still be full of #REF! / #DIV/0! errors
that only surface when Excel opens it. This script forces a recalc using
the `formulas` package (pip-installable, pure Python, no system deps), then
scans every cell for error values.

Output (stdout):
    {
      "success": true,
      "path": "agent-outputs/NVDA.xlsx",
      "total_formulas": 87,
      "errors": []
    }

If errors are found:
    {
      "success": false,
      "path": "agent-outputs/NVDA.xlsx",
      "total_formulas": 87,
      "errors": [
        {"sheet": "Fundamentals", "cell": "D5", "error_type": "#DIV/0!"},
        ...
      ]
    }

Exits 0 on success (no errors), 1 if any error cells were found, 2 on
infrastructure failure (formulas missing, file not found, etc.).
"""

import json
import re
import sys
from pathlib import Path

ERROR_VALUES = {
    "#REF!",
    "#DIV/0!",
    "#VALUE!",
    "#N/A",
    "#NAME?",
    "#NUM!",
    "#NULL!",
    "#GETTING_DATA",
    "#SPILL!",
    "#CALC!",
}


# Formulas package emits cell keys like:
#     '[NVDA.xlsx]FUNDAMENTALS'!D5
# We need to pull out the sheet name and the cell coordinate.
# Formulas keys come quoted: "'[NVDA.xlsx]FUNDAMENTALS'!B6"
CELL_KEY_RE = re.compile(r"\[[^\]]+\]([^'!]+)'?!([A-Z]+\d+)$")


def parse_cell_key(key: str) -> tuple[str, str] | None:
    m = CELL_KEY_RE.search(key)
    if not m:
        return None
    return m.group(1), m.group(2)


def extract_value(ranges_obj) -> str | float | int | None:
    """The formulas library returns Ranges objects whose value is a 2D array.
    For a single cell that's [[value]]. We want the inner scalar.
    """
    try:
        arr = ranges_obj.value
    except AttributeError:
        return ranges_obj
    if arr is None:
        return None
    # numpy 2D array, list of lists, etc.
    try:
        flat = arr[0][0]
    except (IndexError, TypeError):
        return arr
    # numpy scalars stringify cleanly via str()
    return flat


def scan_with_formulas(path: Path) -> tuple[int, list[dict]]:
    try:
        import formulas
    except ImportError:
        raise RuntimeError(
            "The `formulas` package is not installed in the project venv. "
            "Run `./.venv/bin/pip install -r .claude/skills/ensure-deps/requirements.txt`."
        )

    try:
        from openpyxl import load_workbook
    except ImportError:
        raise RuntimeError(
            "openpyxl is not installed in the project venv. "
            "Run `./.venv/bin/pip install -r .claude/skills/ensure-deps/requirements.txt`."
        )

    # Count formulas (the string forms) for the report.
    wb = load_workbook(path, data_only=False)
    formula_count = 0
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        for row in ws.iter_rows():
            for cell in row:
                if isinstance(cell.value, str) and cell.value.startswith("="):
                    formula_count += 1

    # Compile + evaluate every formula in the workbook.
    xl = formulas.ExcelModel().loads(str(path)).finish()
    solution = xl.calculate()

    errors: list[dict] = []
    for key, value in solution.items():
        scalar = extract_value(value)
        # formulas returns XlError objects (which str() to "#DIV/0!" etc.)
        # for error cells, and plain strings for legit string cells.
        scalar_str = str(scalar) if scalar is not None else ""
        if scalar_str in ERROR_VALUES:
            parsed = parse_cell_key(key)
            if not parsed:
                continue
            sheet, coord = parsed
            errors.append({"sheet": sheet, "cell": coord, "error_type": scalar_str})

    return formula_count, errors


def main() -> int:
    if len(sys.argv) != 2:
        print("usage: recalc.py path/to/workbook.xlsx", file=sys.stderr)
        return 2

    path = Path(sys.argv[1]).resolve()
    if not path.exists():
        print(json.dumps({"success": False, "error": f"file not found: {path}"}))
        return 2

    try:
        total_formulas, errors = scan_with_formulas(path)
    except Exception as exc:  # noqa: BLE001
        print(json.dumps({"success": False, "error": "recalc_failed", "detail": str(exc)}))
        return 2

    output = {
        "success": len(errors) == 0,
        "path": str(path.relative_to(Path.cwd())) if path.is_relative_to(Path.cwd()) else str(path),
        "total_formulas": total_formulas,
        "errors": errors,
    }
    print(json.dumps(output, indent=2))
    return 0 if output["success"] else 1


if __name__ == "__main__":
    sys.exit(main())
